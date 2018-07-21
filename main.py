from openpyxl import *

def get_codes (country_code, codes):
    code_list = []

    for code in codes.split(','):
        if code:
            full_code = str(country_code) + str(code.strip())
            code_list.append(full_code)
        else:
            code_list.append(country_code)
    return code_list

def get_breakout (country, destination):
    if destination:
        breakout = str(country) + ' ' + str(destination)
        return breakout
    else:
         breakout = str(country)
         return breakout

def get_destination (breakout, codes, price, date, comment):
    destination = {'breakout': '', 'codes': '', 'price': '', 'date': '', 'comment': ''}
    destination['breakout'] = breakout
    destination['codes'] = codes
    destination['price'] = price
    destination['date'] = date
    if comment:
        destination['comment'] = comment
    else:
        destination['comment'] = ''
    return destination

def get_rows(ws, cell_pos):
    for row in ws.iter_rows(min_row = cell_pos['first_row']+1, max_col = cell_pos['last_col'], max_row = cell_pos['last_row']):
        destination_list.append (get_destination (get_breakout (row[cell_pos['country']].value, row[cell_pos['destination']].value),
                          get_codes (row[cell_pos['country_code']].value, row[cell_pos['codes']].value),
                         row[cell_pos['price']].value, row[cell_pos['date']].value, row[cell_pos['comment']].value))
    
def create_headers(ws):
    ws['A1'] = "Breakout"
    ws['B1'] = "Codes"
    ws['C1'] = "Price"
    ws['D1'] = "Date"
    ws['E1'] = "Comments"
    
def fill_file (ws, destination_list):
    row_number = 2
    sum = len(destination_list)
    for destination in destination_list:
        for code in destination['codes']:
            ws[row_number][0].value = destination['breakout']
            ws[row_number][1].value = code
            ws[row_number][2].value = destination['price']
            ws[row_number][3].value = destination['date']
            ws[row_number][4].value = destination['comment']
            row_number += 1
            print 'Current destination: {0:30} Current code: {1:10}\r'.format(destination['breakout'],code),
            
def get_pos (ws):
    row_number = 0
    cell_pos = {}
    for row in ws.iter_rows(min_row = 1, max_col = 1, max_row = ws.max_row):
        row_number +=1
        if row[0].value == 'Country':
            cell_pos['first_row'] = row_number
            break
    for row in ws.iter_rows(min_row = cell_pos['first_row']+1, max_col = 1, max_row = ws.max_row):
        row_number +=1
        if row[0].value == '':
            cell_pos['last_row'] = row_number-1
            break
    get_cells (ws, cell_pos['first_row'], cell_pos)
    return cell_pos

def get_cells (ws, row_number, cell_pos):
    row = ws[row_number]
    pos = 0
    for cell in row:
        if cell.value == 'Country':
            cell_pos['country'] = pos
        elif cell.value == 'Destination':
            cell_pos['destination'] = pos
        elif cell.value == 'Country Code(s)':
            cell_pos['country_code'] = pos
        elif cell.value == 'City Code(s)':
            cell_pos['codes'] = pos
        elif cell.value == 'Price($)':
            cell_pos['price'] = pos
        elif cell.value == 'Effective Date':
            cell_pos['date'] = pos
        elif cell.value == 'Comments':
            cell_pos['comment'] = pos
        elif cell.value == 'Modification':
            cell_pos['modification'] = pos
        pos +=1
    cell_pos['last_col'] = pos
    return cell_pos

def fnd(lst, key, value):
    for i, dic in enumerate(lst):
        if dic[key] == value:
            return i          
    
def change_codes (wb):
    ws = wb['Code Changes']
    cell_pos = get_pos (ws)
    for row in ws.iter_rows(min_row = cell_pos['first_row'], max_col = cell_pos['last_col'], max_row = cell_pos['last_row']):
        dst = get_breakout (row[cell_pos['country']].value, row[cell_pos['destination']].value)
        index = fnd(destination_list, 'breakout', dst)
        codes = get_codes (row[cell_pos['country_code']].value,row[cell_pos['modification']].value)
        if row[cell_pos['comment']].value == 'Code Added':
            for code in codes:
                if  code in destination_list[index]['codes']:
                    continue
                else:
                    destination_list[index]['codes'].append(code)
        if row[cell_pos['comment']].value == 'Code Removed':
            for code in codes:
                if  code in destination_list[index]['codes']:
                    destination_list[index]['codes'].remove(code)
                else:
                    continue
        if row[cell_pos['comment']].value == 'Destination Removed':
            destination_list.pop(index)

            
original_wb = load_workbook('rates.xlsx')
original_ws = original_wb.active     
destination_list = []
print 'Reading the file...\n'
get_rows(original_ws, get_pos(original_ws))

while True:
    code_changes = raw_input ('Do you have "Code changes list"?(y/n)\n')
    if code_changes == 'y':
        print 'Changing codes...\n'
        change_codes (original_wb)
        break
    elif code_changes == 'n':
        break
    else:
        print ('Please type only "y" or "n"\n')
        continue        
print 'Creating new file...\n'
new_wb = Workbook()
new_ws = new_wb.active
create_headers(new_ws)
fill_file(new_ws, destination_list)
new_wb.save('Edited.xlsx')





